"""
build_nationality_dataset.py

Builds a derived nationality-by-sector dataset from two official Home
Office files:

  - occupation-soc2010-visas-datasets-mar-2024.xlsx (SOC 2010 edition,
    covers 2021 Q1 to 2024 Q1, real nationality data throughout)
  - occupation-soc2020-visas-datasets-mar-2026.xlsx (SOC 2020 edition,
    real nationality data only from 2024 Q4 onward - the period before
    that is suppressed in this file, per the Home Office's own Note 1)

Both files use the same Industry classification (checked directly -
same 22 category names in both), so the Industry field is what's used
here, not the occupation/SOC code, which did change between editions.

2024 Q2 and 2024 Q3 are not in either file with real nationality data,
so they are genuinely missing here too - nothing is estimated, filled,
or interpolated for those two quarters.
"""

import openpyxl
import csv

SECTOR_MAP = {
    "Information and Communications": "Technology",
    "Human Health and Social Work Activities": "Healthcare",
    "Financial and Insurance Activities": "Finance",
    "Construction": "Engineering",
    "Education": "Education",
}


def extract_soc2010(filename):
    """Reads the SOC 2010 file - has an extra 'Occupation' and 'SOC code'
    column compared to the SOC 2020 file, so the column positions differ."""
    wb = openpyxl.load_workbook(filename, data_only=True, read_only=True)
    ws = wb["Data_Occ_D02"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        year, quarter, nationality, region, visa_type, subgroup, occupation, industry, soc_code, major, submajor, minor, unit, grants = row
        if quarter is None:
            continue
        sector = SECTOR_MAP.get(industry)
        if sector is None:
            continue
        rows.append({
            "Year": year, "Quarter": quarter, "Nationality": nationality,
            "Industry": industry, "Sector": sector, "Grants": grants,
            "Source_Dataset": "SOC 2010",
        })
    return rows


def extract_soc2020(filename):
    """Reads the SOC 2020 file - different column order (no separate
    'Occupation' text column, and Industry comes before the SOC groups)."""
    wb = openpyxl.load_workbook(filename, data_only=True, read_only=True)
    ws = wb["Data_Occ_D02"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:
            continue
        year, quarter, nationality, region, visa_type, subgroup, industry, major, submajor, minor, unit, grants = row
        if quarter is None:
            continue
        # skipping suppressed rows entirely here - this is the real,
        # honest way the gap shows up, not something filled in later
        if nationality == "*":
            continue
        sector = SECTOR_MAP.get(industry)
        if sector is None:
            continue
        rows.append({
            "Year": year, "Quarter": quarter, "Nationality": nationality,
            "Industry": industry, "Sector": sector, "Grants": grants,
            "Source_Dataset": "SOC 2020",
        })
    return rows


if __name__ == "__main__":
    print("extracting SOC 2010 file...")
    rows_2010 = extract_soc2010("occupation-soc2010-visas-datasets-mar-2024.xlsx")
    print(f"  {len(rows_2010):,} occupation-level rows in the 5 tracked sectors")

    print("extracting SOC 2020 file...")
    rows_2020 = extract_soc2020("occupation-soc2020-visas-datasets-mar-2026.xlsx")
    print(f"  {len(rows_2020):,} occupation-level rows in the 5 tracked sectors")

    all_rows = rows_2010 + rows_2020
    print(f"combined before aggregation: {len(all_rows):,} rows")

    # the source files have one row per occupation (SOC unit group) within
    # each Year/Quarter/Nationality/Industry combination - collapsing that
    # down to one row per Year/Quarter/Nationality/Industry by summing
    # Grants, since the feature needs a nationality-by-sector total, not
    # individual occupation detail. This is summing real numbers within
    # the same real category, not estimating or filling anything. Industry
    # is kept alongside Sector, as each Sector maps to exactly one
    # Industry string, so this doesn't change the grouping.
    import pandas as pd
    df = pd.DataFrame(all_rows)
    aggregated = df.groupby(
        ["Year", "Quarter", "Nationality", "Industry", "Sector", "Source_Dataset"], as_index=False
    )["Grants"].sum()

    print(f"after aggregating occupation-level rows up to Nationality/Sector: {len(aggregated):,} rows")

    aggregated.to_csv("Nationality_By_Sector.csv", index=False)
    print("saved Nationality_By_Sector.csv")
